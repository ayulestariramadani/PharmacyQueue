import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QMessageBox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class ExcelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Excel Sheet Creator')
        self.setGeometry(100, 100, 300, 200)

        layout = QVBoxLayout()

        self.create_btn = QPushButton('Create Excel Sheet', self)
        self.create_btn.clicked.connect(self.create_excel_sheet)
        layout.addWidget(self.create_btn)

        self.setLayout(layout)

    def create_excel_sheet(self):
        # Get the current month and date
        now = datetime.now()
        month_str = now.strftime('%B_%Y')
        today_str = now.strftime('%d')

        # Define the Excel file name based on the current month
        file_name = f'{month_str}.xlsx'

        # Check if the Excel file exists
        if not os.path.exists(file_name):
            # Create a new Excel workbook and save it
            wb = Workbook()
            wb.save(file_name)

        # Load the workbook
        wb = load_workbook(file_name)

        # Delete the default sheet if it exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Check if the sheet for today's date already exists
        if today_str not in wb.sheetnames:
            # Create a new sheet with today's date
            wb.create_sheet(today_str)

            # Save the workbook
            wb.save(file_name)

            QMessageBox.information(self, 'Success', f'Sheet "{today_str}" created in "{file_name}".')
        else:
            QMessageBox.information(self, 'Info', f'Sheet "{today_str}" already exists in "{file_name}".')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelApp()
    ex.show()
    sys.exit(app.exec_())
